import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from templates.base import (
    setup_sheet, style_header_row, style_data_row, style_total_row,
    FONT_NAME, HEADER_BOLD, PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
    font_title, font_header, font_subheader, font_body, font_caption,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text,
    COLUMN_WIDTHS, ROW_HEIGHTS, FORMATS,
)

wb = Workbook()

# ============================================================
# MESES
# ============================================================
MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
MESES_FULL = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
              "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

# Currency format for Uruguay (UYU)
UYU_FORMAT = '#,##0'

# ============================================================
# SHEET 1: INGRESOS POR UNIDAD
# ============================================================
ws1 = wb.active
ws1.title = "Ingresos por Unidad"

# 29 units: 27 apartments + 2 locales
UNIDADES = [f"Apto {i}" for i in range(1, 28)] + [f"Local {i}" for i in range(1, 3)]

# Columns: A(margin) | B(Unidad) | C-N(Meses Ene-Dic) | O(Total Anual)
LAST_COL = 15  # O = column 15
HEADER_ROW = 4
DATA_START = 5
NUM_UNITS = len(UNIDADES)  # 29
DATA_END = DATA_START + NUM_UNITS - 1  # row 33
TOTAL_ROW = DATA_END + 1  # row 34

setup_sheet(ws1, title="Ingresos por Unidad — Edificio Uruguay", last_col=LAST_COL)

# Column widths
ws1.column_dimensions["B"].width = 16  # Unidad
for col_idx in range(3, LAST_COL + 1):
    ws1.column_dimensions[get_column_letter(col_idx)].width = 14  # Meses + Total

# Header row
headers = ["Unidad"] + MESES + ["Total Anual"]
for i, h in enumerate(headers):
    col = i + 2  # starts at column B
    cell = ws1.cell(row=HEADER_ROW, column=col, value=h)
style_header_row(ws1, HEADER_ROW, 2, LAST_COL)

# Data rows (empty — for pasting)
for idx, unidad in enumerate(UNIDADES):
    row = DATA_START + idx
    # Unit name
    cell = ws1.cell(row=row, column=2, value=unidad)
    cell.font = font_body()
    cell.alignment = align_text()
    
    # Month cells (empty, formatted for currency)
    for col in range(3, 15):  # C to N (12 months)
        cell = ws1.cell(row=row, column=col)
        cell.number_format = UYU_FORMAT
        cell.alignment = align_number()
    
    # Total Anual formula = SUM of months
    total_col = 15  # O
    cell = ws1.cell(row=row, column=total_col)
    cell.value = f"=SUM(C{row}:N{row})"
    cell.number_format = UYU_FORMAT
    cell.alignment = align_number()
    cell.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=NEUTRAL_900)
    
    # Alternating row style
    fill = fill_data_row(idx)
    for col in range(2, LAST_COL + 1):
        ws1.cell(row=row, column=col).fill = fill
    ws1.row_dimensions[row].height = ROW_HEIGHTS["data"]

# Total row
total_cell = ws1.cell(row=TOTAL_ROW, column=2, value="TOTAL INGRESOS")
total_cell.font = font_subheader()
total_cell.alignment = align_text()

for col in range(3, LAST_COL + 1):
    col_letter = get_column_letter(col)
    cell = ws1.cell(row=TOTAL_ROW, column=col)
    cell.value = f"=SUM({col_letter}{DATA_START}:{col_letter}{DATA_END})"
    cell.number_format = UYU_FORMAT
    cell.alignment = align_number()

style_total_row(ws1, TOTAL_ROW, 2, LAST_COL)

# Freeze panes: freeze Unidad column + header row
ws1.freeze_panes = "C5"

# ============================================================
# SHEET 2: GASTOS DEL EDIFICIO
# ============================================================
ws2 = wb.create_sheet("Gastos del Edificio")

# Expense categories (common for building in Uruguay)
CATEGORIAS_GASTO = [
    "Expensas comunes",
    "Impuesto Inmobiliario",
    "Tributo DGI",
    "Contribución Inmobiliaria",
    "Agua (OSE)",
    "Electricidad (UTE) - Areas comunes",
    "Seguro del edificio",
    "Mantenimiento ascensor",
    "Limpieza",
    "Seguridad / Vigilancia",
    "Reparaciones",
    "Mantenimiento general",
    "Honorarios administrador",
    "Servicio de portero",
    "Fondo de reserva",
    "Otros",
]

LAST_COL2 = 15  # O
HEADER_ROW2 = 4
DATA_START2 = 5
NUM_CATS = len(CATEGORIAS_GASTO)
DATA_END2 = DATA_START2 + NUM_CATS - 1
TOTAL_ROW2 = DATA_END2 + 1

setup_sheet(ws2, title="Gastos del Edificio — Edificio Uruguay", last_col=LAST_COL2)

# Column widths
ws2.column_dimensions["B"].width = 32  # Categoría
for col_idx in range(3, LAST_COL2 + 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 14

# Header row
headers2 = ["Categoría"] + MESES + ["Total Anual"]
for i, h in enumerate(headers2):
    col = i + 2
    cell = ws2.cell(row=HEADER_ROW2, column=col, value=h)
style_header_row(ws2, HEADER_ROW2, 2, LAST_COL2)

# Data rows
for idx, cat in enumerate(CATEGORIAS_GASTO):
    row = DATA_START2 + idx
    cell = ws2.cell(row=row, column=2, value=cat)
    cell.font = font_body()
    cell.alignment = align_text()
    
    for col in range(3, 15):
        cell = ws2.cell(row=row, column=col)
        cell.number_format = UYU_FORMAT
        cell.alignment = align_number()
    
    # Total Anual
    cell = ws2.cell(row=row, column=15)
    cell.value = f"=SUM(C{row}:N{row})"
    cell.number_format = UYU_FORMAT
    cell.alignment = align_number()
    cell.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=NEUTRAL_900)
    
    fill = fill_data_row(idx)
    for col in range(2, LAST_COL2 + 1):
        ws2.cell(row=row, column=col).fill = fill
    ws2.row_dimensions[row].height = ROW_HEIGHTS["data"]

# Total row
total_cell2 = ws2.cell(row=TOTAL_ROW2, column=2, value="TOTAL GASTOS")
total_cell2.font = font_subheader()
total_cell2.alignment = align_text()

for col in range(3, LAST_COL2 + 1):
    col_letter = get_column_letter(col)
    cell = ws2.cell(row=TOTAL_ROW2, column=col)
    cell.value = f"=SUM({col_letter}{DATA_START2}:{col_letter}{DATA_END2})"
    cell.number_format = UYU_FORMAT
    cell.alignment = align_number()

style_total_row(ws2, TOTAL_ROW2, 2, LAST_COL2)

# Freeze panes
ws2.freeze_panes = "C5"

# ============================================================
# SHEET 3: RESUMEN / BALANCE
# ============================================================
ws3 = wb.create_sheet("Resumen Balance")

LAST_COL3 = 15
HEADER_ROW3 = 4
DATA_START3 = 5

setup_sheet(ws3, title="Resumen Balance Anual — Edificio Uruguay", last_col=LAST_COL3)

# Column widths
ws3.column_dimensions["B"].width = 28
for col_idx in range(3, LAST_COL3 + 1):
    ws3.column_dimensions[get_column_letter(col_idx)].width = 14

# Headers
headers3 = ["Concepto"] + MESES + ["Total Anual"]
for i, h in enumerate(headers3):
    col = i + 2
    cell = ws3.cell(row=HEADER_ROW3, column=col, value=h)
style_header_row(ws3, HEADER_ROW3, 2, LAST_COL3)

# Row 1: Total Ingresos (linked from Sheet 1)
row_ing = DATA_START3
ws3.cell(row=row_ing, column=2, value="Total Ingresos").font = font_body()
ws3.cell(row=row_ing, column=2).alignment = align_text()
for col in range(3, 15):
    col_letter = get_column_letter(col)
    cell = ws3.cell(row=row_ing, column=col)
    cell.value = f"='Ingresos por Unidad'!{col_letter}{TOTAL_ROW}"
    cell.number_format = UYU_FORMAT
    cell.alignment = align_number()
# Total Anual ingresos
cell = ws3.cell(row=row_ing, column=15)
cell.value = f"=SUM(C{row_ing}:N{row_ing})"
cell.number_format = UYU_FORMAT
cell.alignment = align_number()
cell.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=ACCENT_POSITIVE)
fill_ing = fill_data_row(0)
for col in range(2, LAST_COL3 + 1):
    ws3.cell(row=row_ing, column=col).fill = fill_ing
ws3.row_dimensions[row_ing].height = ROW_HEIGHTS["data"]

# Row 2: Total Gastos (linked from Sheet 2)
row_gas = DATA_START3 + 1
ws3.cell(row=row_gas, column=2, value="Total Gastos").font = font_body()
ws3.cell(row=row_gas, column=2).alignment = align_text()
for col in range(3, 15):
    col_letter = get_column_letter(col)
    cell = ws3.cell(row=row_gas, column=col)
    cell.value = f"='Gastos del Edificio'!{col_letter}{TOTAL_ROW2}"
    cell.number_format = UYU_FORMAT
    cell.alignment = align_number()
cell = ws3.cell(row=row_gas, column=15)
cell.value = f"=SUM(C{row_gas}:N{row_gas})"
cell.number_format = UYU_FORMAT
cell.alignment = align_number()
cell.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=ACCENT_NEGATIVE)
fill_gas = fill_data_row(1)
for col in range(2, LAST_COL3 + 1):
    ws3.cell(row=row_gas, column=col).fill = fill_gas
ws3.row_dimensions[row_gas].height = ROW_HEIGHTS["data"]

# Row 3: Balance (Ingresos - Gastos)
row_bal = DATA_START3 + 2
ws3.cell(row=row_bal, column=2, value="BALANCE").font = font_subheader()
ws3.cell(row=row_bal, column=2).alignment = align_text()
for col in range(3, 16):
    col_letter = get_column_letter(col)
    cell = ws3.cell(row=row_bal, column=col)
    cell.value = f"={col_letter}{row_ing}-{col_letter}{row_gas}"
    cell.number_format = UYU_FORMAT
    cell.alignment = align_number()
    cell.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PRIMARY)
style_total_row(ws3, row_bal, 2, LAST_COL3)

# Row 4: spacer
row_spacer = DATA_START3 + 3
ws3.row_dimensions[row_spacer].height = 12

# Row 5: Desglose por unidad - what each apt should pay monthly (proportional)
row_header2 = DATA_START3 + 4
ws3.cell(row=row_header2, column=2, value="Desglose proporcional por unidad").font = font_title()
ws3.cell(row=row_header2, column=2).alignment = align_title()
ws3.merge_cells(start_row=row_header2, start_column=2, end_row=row_header2, end_column=LAST_COL3)
ws3.row_dimensions[row_header2].height = ROW_HEIGHTS["title"]

# Sub-header row
row_subhead = row_header2 + 1
sub_headers = ["Unidad"] + MESES + ["Total Anual"]
for i, h in enumerate(sub_headers):
    col = i + 2
    cell = ws3.cell(row=row_subhead, column=col, value=h)
style_header_row(ws3, row_subhead, 2, LAST_COL3)

# Each unit's proportional share = Total Gastos / 29 (equal share)
# User can adjust this later if some units pay different proportions
row_unit_start = row_subhead + 1
for idx, unidad in enumerate(UNIDADES):
    row = row_unit_start + idx
    ws3.cell(row=row, column=2, value=unidad).font = font_body()
    ws3.cell(row=row, column=2).alignment = align_text()
    
    for col in range(3, 15):
        col_letter = get_column_letter(col)
        cell = ws3.cell(row=row, column=col)
        # Equal share: Total Gastos / 29 units
        cell.value = f"=IFERROR({col_letter}{row_gas}/{NUM_UNITS},0)"
        cell.number_format = UYU_FORMAT
        cell.alignment = align_number()
    
    # Total Anual for this unit
    cell = ws3.cell(row=row, column=15)
    cell.value = f"=SUM(C{row}:N{row})"
    cell.number_format = UYU_FORMAT
    cell.alignment = align_number()
    
    fill = fill_data_row(idx)
    for col in range(2, LAST_COL3 + 1):
        ws3.cell(row=row, column=col).fill = fill
    ws3.row_dimensions[row].height = ROW_HEIGHTS["data"]

# Total verification row
row_verif = row_unit_start + NUM_UNITS
ws3.cell(row=row_verif, column=2, value="Verificación (suma)").font = font_subheader()
ws3.cell(row=row_verif, column=2).alignment = align_text()
for col in range(3, LAST_COL3 + 1):
    col_letter = get_column_letter(col)
    cell = ws3.cell(row=row_verif, column=col)
    cell.value = f"=SUM({col_letter}{row_unit_start}:{col_letter}{row_verif - 1})"
    cell.number_format = UYU_FORMAT
    cell.alignment = align_number()
style_total_row(ws3, row_verif, 2, LAST_COL3)

# Freeze panes
ws3.freeze_panes = "C5"

# ============================================================
# SAVE
# ============================================================
OUTPUT_PATH = "/home/z/my-project/download/Edificio_Uruguay_Balance.xlsx"
wb.save(OUTPUT_PATH)
print(f"Excel saved to: {OUTPUT_PATH}")
