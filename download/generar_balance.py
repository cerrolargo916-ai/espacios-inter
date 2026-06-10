import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

# ===========================
# READ SOURCE FILES
# ===========================

# --- File 1: Balance definition (ENERO 2026) ---
wb1 = openpyxl.load_workbook('upload/para ia definicion balance.xlsx', data_only=True)
ws1 = wb1['Hoja1']

# Parse apartments from file 1
apartments = {}
for row in ws1.iter_rows(min_row=3, max_row=31, values_only=False):
    contrato = row[0].value  # A: ANDA, PORTO, CGN, DEPOSITO, PROPIEDAD, PORTERIA
    nro = row[1].value       # B: apartment number
    alquiler = row[2].value or 0  # C
    imm = row[3].value or 0       # D
    gc = row[4].value or 0        # E
    san = row[5].value or 0       # F
    otros_merc = row[6].value or 0 # G
    com_san_merc = row[7].value or 0  # H
    iva_san_merc = row[8].value or 0  # I
    com_imm = row[9].value or 0    # J
    iva_imm = row[10].value or 0   # K
    com_lars = row[11].value or 0  # L
    com_anda_conf = row[12].value or 0  # M
    iva_comisiones = row[13].value or 0 # N
    com_gc = row[14].value or 0    # O
    iva_gc = row[15].value or 0    # P
    irpf_rate = row[16].value or 0 # Q (0.105 = 10.5%)
    
    if nro and contrato:
        apto_key = str(nro)
        # Determine unit type
        nro_int = int(nro) if isinstance(nro, (int, float)) else 0
        if nro_int >= 500:
            tipo = 'Local/Depósito'
        elif nro_int >= 400:
            tipo = '4° Piso'
        elif nro_int >= 300:
            tipo = '3° Piso'
        elif nro_int >= 200:
            tipo = '2° Piso'
        elif nro_int >= 100:
            tipo = '1° Piso'
        else:
            tipo = 'PB/Local'
        
        # Build apartment name
        if contrato == 'PORTERIA':
            apto_name = 'Portería'
        elif contrato == 'PROPIEDAD':
            apto_name = f'Propiedad ({nro})'
        else:
            apto_name = f'Apto {nro}'
        
        apartments[apto_key] = {
            'contrato': contrato,
            'nro': nro,
            'tipo': tipo,
            'apto_name': apto_name,
            'alquiler': alquiler,
            'imm': imm,
            'gc': gc,
            'san': san,
            'otros_merc': otros_merc,
            'com_san_merc': com_san_merc,
            'iva_san_merc': iva_san_merc,
            'com_imm': com_imm,
            'iva_imm': iva_imm,
            'com_lars': com_lars,
            'com_anda_conf': com_anda_conf,
            'iva_comisiones': iva_comisiones,
            'com_gc': com_gc,
            'iva_gc': iva_gc,
            'irpf_rate': irpf_rate,
            'retencion_irpf': round(alquiler * irpf_rate, 2) if irpf_rate else 0,
        }

# Summary data from file 1
summary = {
    'saldo_inicial': 0, 'ing_alquileres': 0, 'reembolsos': 0,
    'total_ingresos': 0, 'egresos': 0, 'retiro': 0, 'total_egresos': 0,
    'saldo_final': 0, 'comisiones_lars': 0, 'comisiones_anda': 0,
    'iva_comisiones': 0, 'comisiones_gc': 0, 'iva_gc': 0,
}
for row in ws1.iter_rows(min_row=32, max_row=49, values_only=False):
    a = str(row[0].value or '').strip()
    b = str(row[1].value or '').strip()
    c = row[2].value or 0
    if a == 'SALDO INICIAL': summary['saldo_inicial'] = c
    elif a == 'ING.ALQUILERES': summary['ing_alquileres'] = c
    elif b == 'TOTAL' and not a: summary['total_ingresos'] = c
    elif a == 'EGRESOS': summary['egresos'] = c
    elif a == 'RETIRO': summary['retiro'] = c
    elif a == 'SALDO FINAL': summary['saldo_final'] = c

# --- File 2: Bank statement (FEBRERO 2026) ---
wb2 = openpyxl.load_workbook('upload/lisstado para desglosar.xlsx', data_only=True)
ws2 = wb2['Hoja1']

# Parse bank statement - group by apartment
aptos_bank = {}  # {apto_ref: [{date, concept, debit, credit, balance}...]}
gastos_generales = []  # expenses not tied to a specific apartment

def extract_apto_from_ref(ref):
    """Extract apartment number from reference like 'CERRO LARGO 916/102'"""
    if not ref:
        return None
    m = re.search(r'916/(\d+)', str(ref))
    if m:
        return m.group(1)
    m = re.search(r'918/(\d+)', str(ref))
    if m:
        return f'918-{m.group(1)}'
    m = re.search(r'914/(\w+)', str(ref))
    if m:
        return f'914-{m.group(1)}'
    return None

def extract_apto_from_concept(concept):
    """Try to extract apartment from concept text"""
    if not concept:
        return None
    m = re.search(r'C LARGO (\d+)P', str(concept))
    if m:
        return m.group(1)
    m = re.search(r'916/(\d+)', str(concept))
    if m:
        return m.group(1)
    return None

for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=False):
    date = row[0].value
    ref = row[1].value  # B
    concept = str(row[2].value or '').strip()  # C
    ref_date = row[3].value  # D
    note = str(row[4].value or '').strip()  # E
    debit = row[5].value or 0   # F (egreso)
    credit = row[6].value or 0  # G (ingreso)
    balance = row[7].value or 0 # H
    
    entry = {
        'date': date,
        'ref': ref,
        'concept': concept,
        'ref_date': ref_date,
        'note': note,
        'debit': float(debit),
        'credit': float(credit),
        'balance': float(balance),
    }
    
    # Try to find apartment
    apto = extract_apto_from_ref(ref)
    if not apto:
        apto = extract_apto_from_concept(concept)
    
    # Classify entries
    is_income = concept in ['ALQUILER', 'GASTOS COMUNES', 'Cobro por ANDA', 'Cobro por CONTA'] or credit > 0
    is_expense = any(kw in concept for kw in ['Comisión', 'I.V.A.', 'Retención', 'Ret. IRPF', 'COM.IVA.', 'G/C', 'TARIFA', 'ANTEL', 'UTE', 'FR ED', 'T/D', 'COMISION CONTRATO', 'ADIC.', 'DEPOSITO', 'TRIBUTOS'])
    
    if apto and is_income:
        if apto not in aptos_bank:
            aptos_bank[apto] = {'ingresos': [], 'gastos': []}
        aptos_bank[apto]['ingresos'].append(entry)
    elif apto and is_expense:
        if apto not in aptos_bank:
            aptos_bank[apto] = {'ingresos': [], 'gastos': []}
        aptos_bank[apto]['gastos'].append(entry)
    elif not apto and is_expense:
        gastos_generales.append(entry)
    elif apto:
        if apto not in aptos_bank:
            aptos_bank[apto] = {'ingresos': [], 'gastos': []}
        # Classify by credit/debit
        if credit > 0:
            aptos_bank[apto]['ingresos'].append(entry)
        else:
            aptos_bank[apto]['gastos'].append(entry)
    else:
        gastos_generales.append(entry)

# ===========================
# CREATE OUTPUT EXCEL
# ===========================

wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='0D7377', end_color='0D7377', fill_type='solid')
subheader_font = Font(name='Calibri', bold=True, size=10, color='1A5C5E')
subheader_fill = PatternFill(start_color='E0F2F1', end_color='E0F2F1', fill_type='solid')
total_font = Font(name='Calibri', bold=True, size=11, color='0D7377')
total_fill = PatternFill(start_color='B2DFDB', end_color='B2DFDB', fill_type='solid')
title_font = Font(name='Calibri', bold=True, size=14, color='0D7377')
subtitle_font = Font(name='Calibri', bold=True, size=11, color='546E7A')
data_font = Font(name='Calibri', size=10)
number_format = '#,##0.00'
thin_border = Border(
    left=Side(style='thin', color='B0BEC5'),
    right=Side(style='thin', color='B0BEC5'),
    top=Side(style='thin', color='B0BEC5'),
    bottom=Side(style='thin', color='B0BEC5'),
)
bottom_border = Border(bottom=Side(style='medium', color='0D7377'))

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_cell(ws, row, col, is_number=False):
    cell = ws.cell(row=row, column=col)
    cell.font = data_font
    cell.border = thin_border
    if is_number:
        cell.number_format = number_format
        cell.alignment = Alignment(horizontal='right')
    else:
        cell.alignment = Alignment(horizontal='left')

def style_total_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border

# ===========================
# SHEET 1: INGRESOS POR APARTAMENTO
# ===========================
ws_ing = wb.active
ws_ing.title = 'Ingresos por Apartamento'

# Title
ws_ing.merge_cells('A1:L1')
ws_ing.cell(row=1, column=1, value='EDIFICIO CERRO LARGO 916 - INGRESOS POR UNIDAD').font = title_font
ws_ing.cell(row=1, column=1).alignment = Alignment(horizontal='center')

ws_ing.merge_cells('A2:L2')
ws_ing.cell(row=2, column=1, value='Enero 2026 - Definición Balance').font = subtitle_font
ws_ing.cell(row=2, column=1).alignment = Alignment(horizontal='center')

# Headers
headers_ing = [
    'Nº', 'Unidad', 'Contrato', 'Tipo',
    'Alquiler', 'Gastos Comunes', 'IMM', 'Saneamiento',
    'Otros/Merc', 'Total Ingreso', 'Retención IRPF (10.5%)', 'Ingreso Neto'
]
row_num = 4
for col, h in enumerate(headers_ing, 1):
    ws_ing.cell(row=row_num, column=col, value=h)
style_header_row(ws_ing, row_num, len(headers_ing))

# Data rows - sorted by apartment number
def sort_key(item):
    nro = item[1]['nro']
    try:
        return int(float(nro))
    except:
        return 9999

sorted_apts = sorted(apartments.items(), key=sort_key)

row_num = 5
for idx, (key, apt) in enumerate(sorted_apts, 1):
    total_ingreso = apt['alquiler'] + apt['gc'] + apt['imm'] + apt['san'] + apt['otros_merc']
    ingreso_neto = total_ingreso - apt['retencion_irpf']
    
    ws_ing.cell(row=row_num, column=1, value=idx)
    ws_ing.cell(row=row_num, column=2, value=apt['apto_name'])
    ws_ing.cell(row=row_num, column=3, value=apt['contrato'])
    ws_ing.cell(row=row_num, column=4, value=apt['tipo'])
    ws_ing.cell(row=row_num, column=5, value=apt['alquiler'])
    ws_ing.cell(row=row_num, column=6, value=apt['gc'])
    ws_ing.cell(row=row_num, column=7, value=apt['imm'])
    ws_ing.cell(row=row_num, column=8, value=apt['san'])
    ws_ing.cell(row=row_num, column=9, value=apt['otros_merc'])
    ws_ing.cell(row=row_num, column=10, value=total_ingreso)
    ws_ing.cell(row=row_num, column=11, value=apt['retencion_irpf'])
    ws_ing.cell(row=row_num, column=12, value=ingreso_neto)
    
    for col in range(1, 13):
        style_data_cell(ws_ing, row_num, col, is_number=(col >= 5))
    
    # Alternate row color
    if idx % 2 == 0:
        for col in range(1, 13):
            ws_ing.cell(row=row_num, column=col).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    row_num += 1

# Total row
total_row = row_num
ws_ing.cell(row=total_row, column=4, value='TOTALES')
for col_idx in [5, 6, 7, 8, 9, 10, 11, 12]:
    col_letter = get_column_letter(col_idx)
    ws_ing.cell(row=total_row, column=col_idx, value=f'=SUM({col_letter}5:{col_letter}{total_row-1})')
style_total_row(ws_ing, total_row, len(headers_ing))

# Column widths
col_widths = [5, 16, 12, 12, 14, 14, 12, 12, 14, 14, 18, 14]
for i, w in enumerate(col_widths, 1):
    ws_ing.column_dimensions[get_column_letter(i)].width = w

# Summary section
sum_row = total_row + 2
ws_ing.cell(row=sum_row, column=1, value='RESUMEN DEL PERÍODO').font = Font(name='Calibri', bold=True, size=12, color='0D7377')
ws_ing.merge_cells(f'A{sum_row}:D{sum_row}')

summary_data = [
    ('Saldo Inicial', summary['saldo_inicial']),
    ('Total Ingresos por Alquileres', summary['ing_alquileres']),
    ('Total Egresos', summary['egresos']),
    ('Retiro Propietario', summary['retiro']),
    ('Saldo Final', summary['saldo_final']),
]
for i, (label, val) in enumerate(summary_data):
    r = sum_row + 1 + i
    ws_ing.cell(row=r, column=1, value=label).font = Font(name='Calibri', bold=True, size=10)
    ws_ing.cell(row=r, column=5, value=val).number_format = number_format
    ws_ing.cell(row=r, column=5).font = data_font

# ===========================
# SHEET 2: GASTOS DEL EDIFICIO
# ===========================
ws_gas = wb.create_sheet('Gastos del Edificio')

# Title
ws_gas.merge_cells('A1:H1')
ws_gas.cell(row=1, column=1, value='EDIFICIO CERRO LARGO 916 - GASTOS EN CONJUNTO').font = title_font
ws_gas.cell(row=1, column=1).alignment = Alignment(horizontal='center')

ws_gas.merge_cells('A2:H2')
ws_gas.cell(row=2, column=1, value='Febrero 2026 - Extracto Bancario Desglosado').font = subtitle_font
ws_gas.cell(row=2, column=1).alignment = Alignment(horizontal='center')

# Section A: Comisiones y deducciones por apartamento
ws_gas.merge_cells('A4:H4')
ws_gas.cell(row=4, column=1, value='A) COMISIONES Y DEDUCCIONES POR UNIDAD').font = Font(name='Calibri', bold=True, size=11, color='1A5C5E')
ws_gas.cell(row=4, column=1).fill = subheader_fill

headers_gas = ['Nº', 'Unidad', 'Comisión Adm.', 'I.V.A. Adm.', 'Retención IRPF', 'Com. IVA ANDA', 'Com. IVA CONTA', 'Total Deducciones']
row_num = 5
for col, h in enumerate(headers_gas, 1):
    ws_gas.cell(row=row_num, column=col, value=h)
style_header_row(ws_gas, row_num, len(headers_gas))

# Group deductions by apartment from file 2
apt_deductions = {}
for apto, data in aptos_bank.items():
    com_adm = sum(e['debit'] for e in data['gastos'] if 'Comisión Adm' in e['concept'])
    iva_adm = sum(e['debit'] for e in data['gastos'] if 'I.V.A. Adm' in e['concept'])
    ret_irpf = sum(e['debit'] for e in data['gastos'] if 'Retención IRPF' in e['concept'] or 'Ret. IRPF' in e['concept'])
    com_iva_anda = sum(e['debit'] for e in data['gastos'] if 'COM.IVA. ANDA' in e['concept'])
    com_iva_conta = sum(e['debit'] for e in data['gastos'] if 'COM.IVA. CONTAD' in e['concept'])
    total = com_adm + iva_adm + ret_irpf + com_iva_anda + com_iva_conta
    
    # Get apartment name
    apt_name = f'Apto {apto}'
    for key, apt in apartments.items():
        if str(apt['nro']) == apto:
            apt_name = apt['apto_name']
            break
    
    apt_deductions[apto] = {
        'name': apt_name,
        'com_adm': com_adm,
        'iva_adm': iva_adm,
        'ret_irpf': ret_irpf,
        'com_iva_anda': com_iva_anda,
        'com_iva_conta': com_iva_conta,
        'total': total,
    }

row_num = 6
for idx, (apto, ded) in enumerate(sorted(apt_deductions.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 9999), 1):
    ws_gas.cell(row=row_num, column=1, value=idx)
    ws_gas.cell(row=row_num, column=2, value=ded['name'])
    ws_gas.cell(row=row_num, column=3, value=round(ded['com_adm'], 2))
    ws_gas.cell(row=row_num, column=4, value=round(ded['iva_adm'], 2))
    ws_gas.cell(row=row_num, column=5, value=round(ded['ret_irpf'], 2))
    ws_gas.cell(row=row_num, column=6, value=round(ded['com_iva_anda'], 2))
    ws_gas.cell(row=row_num, column=7, value=round(ded['com_iva_conta'], 2))
    ws_gas.cell(row=row_num, column=8, value=round(ded['total'], 2))
    
    for col in range(1, 9):
        style_data_cell(ws_gas, row_num, col, is_number=(col >= 3))
    if idx % 2 == 0:
        for col in range(1, 9):
            ws_gas.cell(row=row_num, column=col).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    row_num += 1

def sort_key_apto(apto_str):
    try:
        return int(apto_str)
    except:
        return 9999

# Total row for deductions
total_row_a = row_num
ws_gas.cell(row=total_row_a, column=2, value='TOTALES')
for col_idx in [3, 4, 5, 6, 7, 8]:
    col_letter = get_column_letter(col_idx)
    ws_gas.cell(row=total_row_a, column=col_idx, value=f'=SUM({col_letter}6:{col_letter}{total_row_a-1})')
style_total_row(ws_gas, total_row_a, len(headers_gas))

# Section B: Gastos generales del edificio
sec_b_row = total_row_a + 2
ws_gas.merge_cells(f'A{sec_b_row}:H{sec_b_row}')
ws_gas.cell(row=sec_b_row, column=1, value='B) GASTOS GENERALES DEL EDIFICIO').font = Font(name='Calibri', bold=True, size=11, color='1A5C5E')
ws_gas.cell(row=sec_b_row, column=1).fill = subheader_fill

headers_gen = ['Nº', 'Fecha', 'Concepto', 'Referencia', 'Notas', 'Monto', '', '']
row_num = sec_b_row + 1
for col, h in enumerate(headers_gen, 1):
    ws_gas.cell(row=row_num, column=col, value=h)
style_header_row(ws_gas, row_num, 6)

# Categorize general expenses
gastos_categories = {
    'Gastos Comunes (G/C)': [],
    'Tributos Domiciliarios': [],
    'Tarifas Saneamiento': [],
    'UTE / ANTEL': [],
    'Fondo Reparación': [],
    'Comisión Contrato': [],
    'Adicional Mercantil': [],
    'Depósito/Transferencia': [],
    'Otros': [],
}

# Re-process all expenses from file 2
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=False):
    concept = str(row[2].value or '').strip()
    debit = float(row[5].value or 0)
    ref = str(row[1].value or '').strip()
    date = row[0].value
    note = str(row[4].value or '').strip()
    
    if debit == 0:
        continue
    
    # Categorize
    if 'G/C' in concept or 'GASTOS COMUNES' in concept:
        gastos_categories['Gastos Comunes (G/C)'].append((date, concept, ref, note, debit))
    elif 'TRIBUTOS' in concept or 'T/D' in concept:
        gastos_categories['Tributos Domiciliarios'].append((date, concept, ref, note, debit))
    elif 'TARIFA SANEAM' in concept:
        gastos_categories['Tarifas Saneamiento'].append((date, concept, ref, note, debit))
    elif 'UTE' in concept or 'ANTEL' in concept:
        gastos_categories['UTE / ANTEL'].append((date, concept, ref, note, debit))
    elif 'FR ED' in concept:
        gastos_categories['Fondo Reparación'].append((date, concept, ref, note, debit))
    elif 'COMISION CONTRATO' in concept:
        gastos_categories['Comisión Contrato'].append((date, concept, ref, note, debit))
    elif 'ADIC' in concept or 'MERCANTIL' in concept:
        gastos_categories['Adicional Mercantil'].append((date, concept, ref, note, debit))
    elif 'DEPOSITO' in concept:
        gastos_categories['Depósito/Transferencia'].append((date, concept, ref, note, debit))
    elif 'Comisión Adm' in concept or 'I.V.A.' in concept or 'Retención' in concept or 'Ret. IRPF' in concept or 'COM.IVA' in concept:
        continue  # Already counted in Section A
    else:
        gastos_categories['Otros'].append((date, concept, ref, note, debit))

row_num = sec_b_row + 2
item_idx = 1
category_totals = {}

for cat_name, items in gastos_categories.items():
    if not items:
        continue
    
    # Category header
    ws_gas.merge_cells(f'A{row_num}:F{row_num}')
    ws_gas.cell(row=row_num, column=1, value=cat_name).font = subheader_font
    ws_gas.cell(row=row_num, column=1).fill = subheader_fill
    row_num += 1
    
    cat_total = 0
    for date, concept, ref, note, amount in items:
        date_str = date.strftime('%d/%m/%Y') if isinstance(date, datetime) else str(date)
        ws_gas.cell(row=row_num, column=1, value=item_idx)
        ws_gas.cell(row=row_num, column=2, value=date_str)
        ws_gas.cell(row=row_num, column=3, value=concept)
        ws_gas.cell(row=row_num, column=4, value=ref)
        ws_gas.cell(row=row_num, column=5, value=note)
        ws_gas.cell(row=row_num, column=6, value=amount)
        
        for col in range(1, 7):
            style_data_cell(ws_gas, row_num, col, is_number=(col == 6))
        
        cat_total += amount
        item_idx += 1
        row_num += 1
    
    # Category subtotal
    ws_gas.cell(row=row_num, column=5, value=f'Subtotal {cat_name}').font = Font(name='Calibri', bold=True, size=10, italic=True)
    ws_gas.cell(row=row_num, column=6, value=round(cat_total, 2))
    ws_gas.cell(row=row_num, column=6).font = Font(name='Calibri', bold=True, size=10, italic=True, color='0D7377')
    ws_gas.cell(row=row_num, column=6).number_format = number_format
    category_totals[cat_name] = cat_total
    row_num += 1

# Grand total of general expenses
row_num += 1
ws_gas.cell(row=row_num, column=5, value='TOTAL GASTOS GENERALES').font = total_font
ws_gas.cell(row=row_num, column=6, value=round(sum(category_totals.values()), 2))
ws_gas.cell(row=row_num, column=6).font = total_font
ws_gas.cell(row=row_num, column=6).number_format = number_format
style_total_row(ws_gas, row_num, 6)

# Section C: Resumen de gastos
sec_c_row = row_num + 2
ws_gas.merge_cells(f'A{sec_c_row}:F{sec_c_row}')
ws_gas.cell(row=sec_c_row, column=1, value='C) RESUMEN GENERAL DE GASTOS').font = Font(name='Calibri', bold=True, size=11, color='1A5C5E')
ws_gas.cell(row=sec_c_row, column=1).fill = subheader_fill

resumen_headers = ['Categoría', 'Monto']
row_num = sec_c_row + 1
for col, h in enumerate(resumen_headers, 1):
    ws_gas.cell(row=row_num, column=col, value=h)
style_header_row(ws_gas, row_num, 2)

row_num += 1
all_cats = [('Comisiones Administrativas', sum(d['com_adm'] for d in apt_deductions.values())),
            ('I.V.A. sobre Comisiones', sum(d['iva_adm'] for d in apt_deductions.values())),
            ('Retenciones IRPF', sum(d['ret_irpf'] for d in apt_deductions.values())),
            ('Com. IVA ANDA', sum(d['com_iva_anda'] for d in apt_deductions.values())),
            ('Com. IVA CONTA', sum(d['com_iva_conta'] for d in apt_deductions.values()))]
all_cats += [(k, v) for k, v in category_totals.items()]

grand_total = 0
for cat_name, amount in all_cats:
    ws_gas.cell(row=row_num, column=1, value=cat_name)
    ws_gas.cell(row=row_num, column=2, value=round(amount, 2))
    ws_gas.cell(row=row_num, column=2).number_format = number_format
    for col in range(1, 3):
        style_data_cell(ws_gas, row_num, col, is_number=(col == 2))
    grand_total += amount
    row_num += 1

ws_gas.cell(row=row_num, column=1, value='TOTAL GENERAL EGRESOS')
ws_gas.cell(row=row_num, column=2, value=round(grand_total, 2))
style_total_row(ws_gas, row_num, 2)

# Column widths for sheet 2
col_widths_gas = [5, 14, 40, 22, 14, 14, 14, 14]
for i, w in enumerate(col_widths_gas, 1):
    ws_gas.column_dimensions[get_column_letter(i)].width = w

# ===========================
# SHEET 3: DETALLE BANCARIO (bonus - full bank statement)
# ===========================
ws_det = wb.create_sheet('Detalle Bancario Feb 2026')

ws_det.merge_cells('A1:H1')
ws_det.cell(row=1, column=1, value='EDIFICIO CERRO LARGO 916 - EXTRACTO BANCARIO DETALLADO').font = title_font
ws_det.cell(row=1, column=1).alignment = Alignment(horizontal='center')

ws_det.merge_cells('A2:H2')
ws_det.cell(row=2, column=1, value='Febrero 2026').font = subtitle_font
ws_det.cell(row=2, column=1).alignment = Alignment(horizontal='center')

det_headers = ['Fecha', 'Referencia', 'Concepto', 'Fecha Ref.', 'Notas', 'Egreso', 'Ingreso', 'Saldo']
row_num = 4
for col, h in enumerate(det_headers, 1):
    ws_det.cell(row=row_num, column=col, value=h)
style_header_row(ws_det, row_num, len(det_headers))

row_num = 5
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=False):
    date = row[0].value
    ref = row[1].value or ''
    concept = row[2].value or ''
    ref_date = row[3].value
    note = row[4].value or ''
    debit = row[5].value or 0
    credit = row[6].value or 0
    balance = row[7].value or 0
    
    date_str = date.strftime('%d/%m/%Y') if isinstance(date, datetime) else str(date)
    ref_date_str = ref_date.strftime('%d/%m/%Y') if isinstance(ref_date, datetime) else str(ref_date or '')
    
    ws_det.cell(row=row_num, column=1, value=date_str)
    ws_det.cell(row=row_num, column=2, value=str(ref))
    ws_det.cell(row=row_num, column=3, value=str(concept))
    ws_det.cell(row=row_num, column=4, value=ref_date_str)
    ws_det.cell(row=row_num, column=5, value=str(note))
    ws_det.cell(row=row_num, column=6, value=float(debit) if debit else 0)
    ws_det.cell(row=row_num, column=7, value=float(credit) if credit else 0)
    ws_det.cell(row=row_num, column=8, value=float(balance) if balance else 0)
    
    for col in range(1, 9):
        style_data_cell(ws_det, row_num, col, is_number=(col >= 6))
    
    if row_num % 2 == 0:
        for col in range(1, 9):
            ws_det.cell(row=row_num, column=col).fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    row_num += 1

# Column widths
col_widths_det = [12, 22, 40, 12, 12, 14, 14, 14]
for i, w in enumerate(col_widths_det, 1):
    ws_det.column_dimensions[get_column_letter(i)].width = w

# ===========================
# SAVE
# ===========================
output_path = '/home/z/my-project/download/Balance_Edificio_Cerro_Largo_916.xlsx'
wb.save(output_path)
print(f'Excel guardado en: {output_path}')
print(f'Planillas: {wb.sheetnames}')
print(f'Apartamentos procesados: {len(apartments)}')
print(f'Categorías de gastos: {len([v for v in category_totals.values() if v > 0])}')
